import random
import numpy
from openpyxl import Workbook, load_workbook

# Python program to translate
# speech to text and text to speech


import speech_recognition as sr
import pyttsx3

# Initialize the recognizer
r = sr.Recognizer()


# Function to convert text to
# speech
def SpeakText(command):
    # Initialize the engine
    engine = pyttsx3.init()
    engine.say(command)
    engine.runAndWait()


# Loop infinitely for user to
# speak

while (1):

    # Exception handling to handle
    # exceptions at the runtime
    try:

        # use the microphone as source for input.
        with sr.Microphone() as source2:

            # wait for a second to let the recognizer
            # adjust the energy threshold based on
            # the surrounding noise level
            r.adjust_for_ambient_noise(source2, duration=0.2)

            # listens for the user's input
            audio2 = r.listen(source2)

            # Using google to recognize audio
            MyText = r.recognize_google(audio2)
            MyText = MyText.lower()

            print("Did you say ", MyText)
            SpeakText(MyText)

    except sr.RequestError as e:
        print("Could not request results; {0}".format(e))

    except sr.UnknownValueError:
        print("unknown error occurred")
# workbook = load_workbook(filename="Saved translations.xlsx")
#
# sheets = workbook.sheetnames
# sheet = workbook.active
# print(f"{sheets = }")
# print(f"{sheet = }")
#
# res = sheet["C1:C50"]
# words = [x[0].value for x in res]
# print(f"{words = }")
#
# res = sheet["D1:D50"]
# translations = [x[0].value for x in res]
# print(f"{translations = }")
#
# list1 = []
#
# # for i in range(len(words)):
# #     list1.append(Words(word=words[i], translation=translations[i]))
#
# for i in range(10):
#     list1.append(Words(word=words[i], translation=translations[i]))
#
#
#
# Words.objects.bulk_create(list1)



# lst1 = [1,2,3,4,5,6,7,8,9,10]
# lst2 = [1,3,1,1,1,1,1,1,1,1]
# lst2 = [x / sum(lst2) for x in lst2]
# print([round(x, 3) for x in lst2])
# count = [0,0,0,0,0,0,0,0,0,0]
# for i in range(10):
#     lst3 = random.choices(lst1, weights=lst2, k=4)
#     lst3 = numpy.random.choice(lst1, p=lst2, size=4,replace=False)
#     for j in lst3:
#         count[j-1] += 1
#     print(f"{list(lst3) = }")
#
# print(lst3[0])
# print(count)






# res = sheet["C1:C50"]
# # res = sheet["B"]
# # print(f"{res = }")
# res = [x[0].value for x in res]
# print(f"{res = }")
#
# res = sheet["C"]
# # print(f"{res = }")
# res = [x.value for x in res]
# print(f"{res = }")
# wb = Workbook()
#
# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
#
# # Save the file
# wb.save("sample.xlsx")
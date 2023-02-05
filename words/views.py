import random
import numpy

from django.contrib.auth import logout, login
from django.contrib.auth.views import LoginView
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.views.generic import CreateView

from rest_framework import generics
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .forms import *
from .models import *
from .serializers import *
from .translations import make_google_translation
from .encryption import do_decrypt

menu = [{'title': 'Main page', 'url_name': 'home'},
        {'title': 'Saved', 'url_name': 'saved'},
        {'title': 'Collections', 'url_name': 'collections'},
        {'title': 'Learn', 'url_name': 'learn'},
        {'title': 'Test', 'url_name': 'test'},
        {'title': 'About', 'url_name': 'about'},
        ]

LEARNING_RATING_DECREMENT = 1
LEARNING_RATING_INCREMENT = 2
NUMBER_OF_WORDS_TO_TEST = 10


def get_user_and_session_id(request, *args, **kwargs):
    """
    Func to be called on any request to check if user is logged in or not,
    providing id for a session/user

    :return: user_and_session_record, user_and_session_id
    """
    if request.session.session_key and User_And_Session.objects.filter(session=request.session.session_key).exists():
        user_and_session_record = User_And_Session.objects.get(session=request.session.session_key)
    else:
        # print("no user_and_session object found, creating new")
        if request.user.is_authenticated:
            if User_And_Session.objects.filter(user=request.user).exists():

                # update table User_And_Session with session data
                User_And_Session.objects.filter(user=request.user).update(
                    session=Session.objects.get(session_key=request.session.session_key)
                )
                user_and_session_record = User_And_Session.objects.get(user=request.user)
            else:
                user_and_session_record = User_And_Session.objects.create(
                    user=User.objects.get(id=request.user.id),
                    session=Session.objects.get(session_key=request.session.session_key)
                )
        else:
            # print(f"not authenticated")

            if not request.session.session_key:
                request.session.create()

            # In case of deleting session
            if not Session.objects.filter(session_key=request.session.session_key).exists():
                request.session.create()

            session_key = request.session.session_key

            if not User_And_Session.objects.filter(session=request.session.session_key).exists():
                user_and_session_record = User_And_Session.objects.create(
                    session=Session.objects.get(session_key=session_key))

            if User_And_Session.objects.filter(session=request.session.session_key).exists():
                user_and_session_record = User_And_Session.objects.get(session=request.session.session_key)

        # clear User_And_Session objects and words without users
        Words.objects.filter(
            user_and_session__in=User_And_Session.objects.filter(session=None).filter(user=None)).delete()
        User_And_Session.objects.filter(session=None).filter(user=None).delete()

    user_and_session_id = user_and_session_record.id
    return user_and_session_record, user_and_session_id


def index(request, *args, **kwargs):
    user_and_session, user_and_session_id = get_user_and_session_id(request)

    word = None
    translation = None
    # form1 = None

    if request.method == 'GET':
        pass

    if request.method == 'POST':
        # if button for translation was pressed
        if request.POST.get('do_translate'):
            word = request.POST.get('word').strip()
            if word:
                # try to find translation in Words_Base
                if Words_Base.objects.filter(word=word).exists():
                    translation = Words_Base.objects.filter(word=word)[0].translation
                    # print(f"translation from base")
                else:
                    translation = make_google_translation(word)['translatedText'].strip()

        # if button for saving word and translation was pressed
        if request.POST.get('do_save'):
            if request.POST.get('word') and request.POST.get('translation'):
                word = request.POST.get('word')
                translation = request.POST.get('translation')
                if word != '' and translation != '':
                    similar_words = list(
                        Words.objects.filter(user_and_session__id=user_and_session_id).filter(
                            word=word))
                    if len(similar_words) == 0:
                        Words.objects.create(
                            word=word,
                            translation=translation,
                            user_and_session=user_and_session,
                        )
                    else:
                        Words.objects.filter(user_and_session__id=user_and_session_id).filter(word=word). \
                            update(translation=translation)

        # if the word was chosen (button pressed)
        if request.POST.get('change_word'):
            record = Words.objects.get(pk=int(request.POST.get('change_word')))
            word = record.word
            translation = record.translation

        # if button for deletion word was pressed
        if request.POST.get('delete_word'):
            Words.objects.filter(pk=request.POST.get('delete_word')).delete()
            word = request.session['last_word']['word']
            translation = request.session['last_word']['translation']

    # limit main page words to show by 6. Pagination left for possible further needs
    words = Words.objects.filter(user_and_session=user_and_session).order_by('-pk')[:6]
    paginator = Paginator(words, 6)
    page_obj = paginator.get_page(request.GET.get('page'))

    form1 = WordTranslationForm({'word': word, 'translation': translation})

    if 'last_word' not in request.session:
        request.session['last_word'] = {}
    request.session['last_word']['word'] = word
    request.session['last_word']['translation'] = translation
    request.session.save()

    context = {
        'menu': menu,
        'title': menu[0]['title'],
        'form1': form1,
        'page_obj': page_obj,
    }
    return render(request, 'words/index.html', context=context)


def saved(request, *args, **kwargs):
    user_and_session, user_and_session_id = get_user_and_session_id(request)
    # print(f"{request.POST = }")
    if request.method == 'POST':
        if request.POST.get('delete_word'):
            Words.objects.filter(pk=request.POST.get('delete_word')).delete()

        if request.POST.get('change_word'):
            return index(request)
            # return redirect('home')

    words_records = Words.objects.filter(user_and_session__id=user_and_session_id).order_by('-pk')

    paginator = Paginator(words_records, 30)
    page_obj = paginator.get_page(request.GET.get('page'))
    context = {
        'menu': menu,
        'title': menu[1]['title'],
        'page_obj': page_obj,
    }
    return render(request, 'words/saved.html', context=context)


def collections(request, *args, **kwargs):
    """
    View to save for learning a list of words (out of the box)
    """
    user_and_session, user_and_session_id = get_user_and_session_id(request)
    words_records = []
    choose_collection_form = Choose_collection_form
    choose_collection = Collection.objects.all()
    collection_name = None

    if request.method == 'POST':
        if request.POST.get('Choose_collection'):
            collection_name = request.POST.get('Choose_collection')
            collection_record = Collection.objects.filter(name=collection_name)
            if collection_record:
                collection_record = collection_record[0]
                words_base_collection_words_base_ids = Words_Base_Collection.objects. \
                    filter(collection=collection_record).values_list('words_base', flat=True)
                words_records = Words_Base.objects.filter(id__in=words_base_collection_words_base_ids)
                # print(f"{words_records = }")

        if request.POST.get('collection_name_to_save'):
            collection_name = request.POST.get('collection_name_to_save')
            collection_record = Collection.objects.filter(name=collection_name)
            if collection_record:
                collection_record = collection_record[0]
                # print(f"{collection_record = }")
                Words_Base_ids = Words_Base_Collection.objects.filter(collection=collection_record). \
                    values_list('words_base', flat=True)
                # print(f"{Words_Base_ids = }")
                list1 = list(Words_Base.objects.filter(pk__in=Words_Base_ids).values_list('word', 'translation'))
                # print(f"{list1 = }")
                words_words = Words.objects.filter(user_and_session=user_and_session).values_list('word', flat=True)
                list2 = [Words(word=x[0], translation=x[1], user_and_session=user_and_session) for x in list1 if
                         x[0] not in words_words]
                # print(f"{list2 = }")
                if len(list2) > 0:
                    Words.objects.bulk_create(list2)
                    return redirect('saved')

    context = {
        'menu': menu,
        'title': menu[2]['title'],
        'choose_collection_form': choose_collection_form,
        'choose_collection': choose_collection,
        'words_records': words_records,
        'collection_name': collection_name,
    }
    return render(request, 'words/collections.html', context=context)


def learn(request, *args, **kwargs):
    user_and_session, user_and_session_id = get_user_and_session_id(request)

    words_pks = list(Words.objects.filter(user_and_session__id=user_and_session_id).filter(learned=False). \
                     values_list('pk', flat=True))

    amount_of_words_to_learn = 1
    if request.method == 'GET':
        amount_of_words_to_learn = min(5, len(words_pks))
    else:
        if 'amount_of_words_to_learn' in request.POST and int(request.POST['amount_of_words_to_learn']) > 1:
            amount_of_words_to_learn = min(int(request.POST['amount_of_words_to_learn']), len(words_pks))

    pk_list_random = random.sample(words_pks, k=amount_of_words_to_learn)
    words_records = list(Words.objects.filter(pk__in=pk_list_random))

    form_input_number_of_words_to_learn = ChooseAmountOfWordsToLearnForm(
        {'amount_of_words_to_learn': amount_of_words_to_learn}
    )

    if 'add_words_to_test' not in request.session:
        request.session['add_words_to_test'] = {}
    request.session['add_words_to_test']['words_pks'] = pk_list_random
    request.session.save()

    context = {
        'menu': menu,
        'title': menu[3]['title'],
        'words_records': words_records,
        'form_input_number_of_words_to_learn': form_input_number_of_words_to_learn,
    }
    return render(request, 'words/learn.html', context=context)


def add_words_from_excel():
    """
    Func for adding words from excel table
    """
    from openpyxl import Workbook, load_workbook

    workbook = load_workbook(filename="words/Saved translations.xlsx")

    sheets = workbook.sheetnames
    sheet = workbook.active
    print(f"{sheets = }")
    print(f"{sheet = }")

    res = sheet["C2:C100"]
    words = [x[0].value for x in res]
    print(f"{words = }")

    res = sheet["D2:D100"]
    translations = [x[0].value for x in res]
    print(f"{translations = }")

    list1 = []

    for i in range(len(words)):
        list1.append(Words_Base(word=words[i], translation=translations[i]))

    # for i in range(10):
    #     list1.append(Words(word=words[i], translation=translations[i]))

    Words_Base.objects.bulk_create(list1)

    list2 = []
    for i in list(Words_Base.objects.all()):
        list2.append(Words_Base_Collection(words_base=i, ))
    Words_Base_Collection.objects.bulk_create(list2)
    # Words_Base.objects.create(word=words[0], translation=translations[0])

    # Words.objects.all().update(learning_rating=10)


def test(request, *args, **kwargs):
    user_and_session, user_and_session_id = get_user_and_session_id(request)
    # add_words_from_excel()

    # if language was changed get language value
    if 'test' in request.session and 'lang' in request.session['test']:
        language = request.session['test']['lang']
    else:
        language = 'EN'

    tested_word = None
    tested_word_pk = None
    tested_word_rating = None
    answer_given = None
    correct_answer_given = None
    renew = False
    current_test_stats = (0, 0)

    # First time create a random list of words
    if request.method == 'GET' or request.POST.get('add_words_to_test'):
        if request.POST.get('add_words_to_test'):
            Words.objects.filter(pk__in=request.session['add_words_to_test']['words_pks']).update(learned=True)
        words_pks = list(Words.objects.filter(user_and_session__id=user_and_session_id).filter(learned=True). \
                         values_list('pk', flat=True))
        if len(words_pks) < NUMBER_OF_WORDS_TO_TEST:
            number_of_words_to_test = len(words_pks)
        else:
            number_of_words_to_test = NUMBER_OF_WORDS_TO_TEST
        pk_list_random = random.sample(words_pks, k=number_of_words_to_test)
        words_records = list(Words.objects.filter(pk__in=pk_list_random))
        random.shuffle(words_records)
        if len(words_records) > 0:
            tested_word_record = words_records[0]
            tested_word = (tested_word_record.word, tested_word_record.translation)
            tested_word_pk = tested_word_record.pk
            tested_word_rating = words_records[0].learning_rating
        random.shuffle(words_records)
        answers_to_choose_records = words_records
        answers_to_choose = [(x.pk, x.word, x.translation) for x in words_records]
        # print(f"{answers_to_choose = }")

    if request.method == 'POST':
        if request.POST.get('change_lang') == 'True':
            if request.session['test']['lang'] == 'EN':
                language = 'RU'
            else:
                language = 'EN'
            tested_word = request.session['test']['tested_word']
            tested_word_pk = request.session['test']['tested_word_pk']
            tested_word_rating = request.session['test']['tested_word_rating']
            answers_to_choose = request.session['test']['answers_to_choose']

        if request.POST.get('chosen_answer'):
            answer_given = True
            # If answer is correct, create a list of words according to learning rating,
            # higher the rating, more often you get this word for testing
            if int(request.POST.get('chosen_answer')) == request.session['test']['tested_word_pk']:
                # print("correct")
                correct_answer_given = True
                if request.session['test']['tested_word_rating'] > 1 + LEARNING_RATING_DECREMENT:
                    record = Words.objects.get(pk=int(request.session['test']['tested_word_pk']))
                    record.learning_rating -= LEARNING_RATING_DECREMENT
                    record.save()

                pks_and_learning_ratings = Words.objects.filter(user_and_session__id=user_and_session_id).filter(
                    learned=True).values_list('pk', 'learning_rating')
                # print(f"{records = }")
                pk_list = [record[0] for record in pks_and_learning_ratings]
                if len(pk_list) > 0:
                    learning_rating_list = [record[1] for record in pks_and_learning_ratings]
                    number_of_words_to_test = NUMBER_OF_WORDS_TO_TEST
                    if len(pk_list) < number_of_words_to_test:
                        number_of_words_to_test = len(pk_list)
                    rates_for_choose = [x / sum(learning_rating_list) for x in learning_rating_list]
                    # this numpy method allows to choose only unique words
                    pk_list_random = numpy.random.choice(pk_list,
                                                         p=rates_for_choose,
                                                         size=number_of_words_to_test,
                                                         replace=False)
                    words_records = list(Words.objects.filter(pk__in=pk_list_random))

                    word = words_records[0]
                    for i in range(len(words_records)):
                        if words_records[i].learning_rating > word.learning_rating:
                            word = words_records[i]

                    tested_word_record = word
                    tested_word = (tested_word_record.word, tested_word_record.translation)
                    tested_word_pk = tested_word_record.pk
                    tested_word_rating = tested_word_record.learning_rating

                    random.shuffle(words_records)
                    answers_to_choose = [(x.pk, x.word, x.translation) for x in words_records]
            else:
                record = Words.objects.get(pk=int(request.session['test']['tested_word_pk']))
                record.learning_rating += LEARNING_RATING_INCREMENT
                record.save()
                tested_word = request.session['test']['tested_word']
                tested_word_pk = request.session['test']['tested_word_pk']
                tested_word_rating = record.learning_rating
                answers_to_choose = request.session['test']['answers_to_choose']
        if request.POST.get('renew') == 'True':
            renew = True
            tested_word = request.session['test']['tested_word']
            tested_word_pk = request.session['test']['tested_word_pk']
            tested_word_rating = request.session['test']['tested_word_rating']
            answers_to_choose = request.session['test']['answers_to_choose']

    # making statistics
    number_of_all_rows_in_db = Words.objects.filter(user_and_session__id=user_and_session_id).count()
    number_of_learned_words_in_db = Words.objects.filter(user_and_session__id=user_and_session_id). \
        filter(learned=True).count()
    number_of_words_with_good_rating = Words.objects.filter(user_and_session__id=user_and_session_id). \
        filter(learning_rating__lt=10).count()
    stats = number_of_all_rows_in_db, number_of_learned_words_in_db, number_of_words_with_good_rating
    if 'test' not in request.session:
        request.session['test'] = {}
    if renew:
        correct_answers, answers = (0, 0)
    else:
        if 'current_test_stats' in request.session['test']:
            correct_answers, answers = request.session['test']['current_test_stats']
        else:
            correct_answers, answers = current_test_stats
        if answer_given:
            answers += 1
        if correct_answer_given:
            correct_answers += 1
    current_test_stats = (correct_answers, answers)

    # saving info to current session (saved in db)
    request.session['test']['answers_to_choose'] = answers_to_choose
    request.session['test']['tested_word_pk'] = tested_word_pk
    request.session['test']['tested_word'] = tested_word
    request.session['test']['tested_word_rating'] = tested_word_rating
    request.session['test']['lang'] = language
    request.session['test']['stats'] = stats
    request.session['test']['current_test_stats'] = current_test_stats
    request.session.save()

    context = {
        'menu': menu,
        'title': menu[4]['title'],
        'word_to_test': tested_word,
        'lang': language,
        'answers_to_choose': answers_to_choose,
        'statistics': stats,
        'tested_word_rating': tested_word_rating,
        'current_test_stats': current_test_stats,
    }
    return render(request, 'words/test.html', context=context)


def about(request, *args, **kwargs):
    context = {
        'menu': menu,
        'title': menu[5]['title'],
    }
    return render(request, 'words/about.html', context=context)


class RegisterUser(CreateView):
    form_class = RegisterUserForm
    template_name = 'words/register.html'
    success_url = reverse_lazy('login')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Sign in'
        context['menu'] = menu
        # print(f"1 {self.request = }")
        # print(f"1 {self.request.session.session_key = }")
        return context

    def form_valid(self, form):
        user = form.save()
        user_and_session_record = User_And_Session.objects.get(session=self.request.session.session_key)
        user_and_session_id = user_and_session_record.id
        login(self.request, user)
        # Here we save new session and logged user to record that contained session before login,
        # so that user_and_session_id will not change after login
        User_And_Session.objects.filter(pk=user_and_session_id). \
            update(session=Session.objects.get(session_key=self.request.session.session_key), user=user)
        return redirect('home')


class LoginUser(LoginView):
    form_class = LoginUserForm
    template_name = 'words/login.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Login'
        context['menu'] = menu
        return context

    def get_success_url(self):
        # print(f"5{self.request = }")
        # print(f"5{self.request.session.session_key = }")
        return reverse_lazy('home')


# only if not debug mode
def pageNotFound(request, exception):
    return redirect('home')


def logout_user(request):
    logout(request)
    return redirect('login')

# def login(request, *args, **kwargs):
#     context = {
#         'menu': menu,
#         'title': menu[-1]['title'],
#     }
#     return render(request, 'words/login.html', context=context)

# def register(request, *args, **kwargs):
#     context = {
#         'menu': menu,
#         'title': menu[-1]['title'],
#     }
#     return render(request, 'words/login.html', context=context)



# class WordsAPIView(generics.ListCreateAPIView):
#     queryset = Words.objects.all()[:5]
#     serializer_class = WordSerializer
#     # permission_classes = (IsAuthenticated,)
#
# class WordsAPIUpdate(generics.UpdateAPIView):
#     queryset = Words.objects.all()
#     serializer_class = WordSerializer
#
# class WordsAPIDelete(generics.DestroyAPIView):
#     queryset = Words.objects.all()
#     serializer_class = WordSerializer

class Translation(APIView):
    def get(self, request):
        # print()
        # print(f"{request = }")
        encrypted_data = request.GET.get('encrypted_data', None)
        # print(f"{encrypted_data = }")
        if encrypted_data:
            data = do_decrypt(encrypted_data)
            # print(f"{data = }")
            word = data['word']
            target_language = data['target_language']
            source_language = data['source_language']

            # print(f"{word = }")

            if word and target_language and source_language:
                # try to find translation in Words_Base
                if Words_Base.objects.filter(word=word).exists():
                    translation = Words_Base.objects.filter(word=word)[0].translation
                    # print(f"translation from base")
                else:
                    translation = make_google_translation(word, target_language, source_language)['translatedText'].strip()
                result = JsonResponse({'word': word, 'translation': translation})
            else:
                result = JsonResponse({'error': 'validation error'})
        else:
            result = JsonResponse({'error': 'token error'})

        # print(f"{result = }")

        return result

    # permission_classes = (IsAuthenticated, )
